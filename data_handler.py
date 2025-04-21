import csv
from openpyxl import Workbook

"""
Function that will append data to a csv file 

def dataToCSV(company, data)

Parameters:
csv_path: CSV Path to the desired location.
company: Company name 
data: Data information of candidate that were collected
"""

def dataToCSV(csv_path, company, data):
    with open(csv_path, "a+", newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        
        file.seek(0, 2)
        if file.tell() == 0:
            writer.writerow(['Company', 'Title', 'Name', 'Link'])

        for name, members in data.members.items():
            for person in members:
                row = [company, person.title, person.name, person.link]
                writer.writerow(row)


"""
Function that converts csv to excel and organizes the data in columns for each company

def csvToExcel()
parameters:
csv_path: CSV Path
xlsx_path: XLSX Path
titles: Job Titles
"""
def csvToExcel(csv_path, xlsx_path, titles):
    wb = Workbook()
    ws = wb.active

    with open(csv_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        headers = next(reader)
        
        col_num = 1 
        row_num = 1
        
        ws.cell(row=row_num, column=col_num).value = 'Company'
        for title in titles:
            col_num += 1
            ws.cell(row=row_num, column=col_num).value = title          

        col_num = 1
        row_num = 2
        company = None 
        
        for row in reader:
            if company != row[0]:
                company = row[0]
                ws.cell(row=row_num, column=1).value = company
                row_num += 1

            company = row[0]
            title = row[1]
            name = row[2]
            profile_link = row[3]

            col_num = titles.index(title) + 2

            cell = ws.cell(row=row_num-1, column=col_num)
            cell.value = name
            cell.hyperlink = profile_link

    wb.save(xlsx_path)